#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel/CSV TUI 工具 - 终端界面只读查看 Excel/CSV 文件
支持: Sheet 选择、单元格浏览、行详情、搜索、过滤
"""

import argparse
import bisect
import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from textual.app import App, ComposeResult

# 可选依赖，xls 格式需要
try:
    import xlrd
    HAS_XLS = True
except ImportError:
    HAS_XLS = False

from textual.binding import Binding
from textual.containers import Vertical, ScrollableContainer
from textual.events import Key, Resize
from textual.message import Message
from textual.screen import Screen
from rich.markup import escape as rich_escape
from textual.widgets import DataTable, Footer, Header, Input, Static


def _displayWidth(s: str) -> int:
    """计算字符串在终端中的显示宽度（CJK=2，英文=1）"""
    w = 0
    for c in s:
        if (
            "\u4e00" <= c <= "\u9fff"
            or "\uff00" <= c <= "\uffef"
            or c in "，。！？、；：""''（）【】"
        ):
            w += 2
        else:
            w += 1
    return w


def _padToDisplayWidth(s: str, width: int, truncate: bool = True) -> str:
    """按显示宽度填充或截断，保证对齐"""
    if truncate and _displayWidth(s) > width:
        result = []
        cur = 0
        for c in s:
            if cur + _displayWidth(c) > width - 2:
                result.append("..")
                break
            cur += _displayWidth(c)
            result.append(c)
        s = "".join(result)
    return s + " " * (width - _displayWidth(s))


def _padToDisplayWidthRight(s: str, width: int) -> str:
    """按显示宽度右对齐填充"""
    return " " * (width - _displayWidth(s)) + s


def _formatDisplayValue(val) -> str:
    """格式化显示值：1.0 -> 1，并移除换行符防止 wrap"""
    if val is None:
        return ""
    s = str(val).strip().replace("\n", " ").replace("\r", " ")
    if not s:
        return ""
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except ValueError:
        pass
    return s


def _escapeForRich(s: str) -> str:
    """转义单元格内容中的 Rich 标记字符（如 {0}、[、] 等），避免 MarkupError"""
    return rich_escape(s)


# ---------- 统一 Sheet 数据结构 ----------

class _XlsCell:
    def __init__(self, sheet: "_XlsSheetWrapper", row: int, col: int) -> None:
        self._sheet = sheet
        self._row = row
        self._col = col

    @property
    def value(self) -> str:
        val = self._sheet._data.get((self._row, self._col), "")
        return "" if val is None else str(val)


class _XlsSheetWrapper:
    def __init__(self, name: str, data: dict[tuple[int, int], str]) -> None:
        self.title = name
        self._data = data
        self._maxRow = max((r for r, _ in data), default=1) if data else 1
        self._maxCol = max((c for _, c in data), default=1) if data else 1

    @property
    def max_row(self) -> int:
        return self._maxRow

    @property
    def max_column(self) -> int:
        return self._maxCol

    def cell(self, row: int, column: int) -> _XlsCell:
        return _XlsCell(self, row, column)


class _XlsWorkbook:
    def __init__(self, sheets: dict[str, _XlsSheetWrapper]) -> None:
        self._sheets = sheets
        self.sheetnames = list(sheets.keys())

    def __getitem__(self, name: str) -> _XlsSheetWrapper:
        return self._sheets[name]


# ---------- 格式加载 ----------

def _loadXls(path: str) -> _XlsWorkbook:
    """加载 xls，使用 row_values 批量读取加速"""
    if not HAS_XLS:
        raise RuntimeError("xls support requires xlrd: pip install xlrd")
    rd = xlrd.open_workbook(path)
    sheets = {}
    for i in range(rd.nsheets):
        sh = rd.sheet_by_index(i)
        data: dict[tuple[int, int], str] = {}
        for r in range(sh.nrows):
            rowVals = sh.row_values(r)
            for c in range(sh.ncols):
                val = rowVals[c] if c < len(rowVals) else ""
                data[(r + 1, c + 1)] = "" if val is None else str(val)
        if not data:
            data[(1, 1)] = ""
        sheets[sh.name] = _XlsSheetWrapper(sh.name, data)
    return _XlsWorkbook(sheets)


def _loadReadOnlyXlsx(path: str) -> _XlsWorkbook:
    """只读模式加载 xlsx，使用 read_only 流式读取大幅加速"""
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            data: dict[tuple[int, int], str] = {}
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    val = cell.value
                    data[(row_idx, col_idx)] = (
                        "" if val is None else str(val).strip().replace("\n", " ").replace("\r", " ")
                    )
            if not data:
                data[(1, 1)] = ""
            sheets[name] = _XlsSheetWrapper(name, data)
        return _XlsWorkbook(sheets)
    finally:
        wb.close()


def _loadCsv(path: str) -> _XlsWorkbook:
    """加载 CSV 文件，自动尝试 utf-8-sig / gbk 编码"""
    sheetName = Path(path).stem
    data: dict[tuple[int, int], str] = {}
    for encoding in ("utf-8-sig", "gbk", "latin-1"):
        try:
            with open(path, newline="", encoding=encoding) as f:
                reader = csv.reader(f)
                for rowIdx, row in enumerate(reader, start=1):
                    for colIdx, val in enumerate(row, start=1):
                        data[(rowIdx, colIdx)] = val.strip().replace("\n", " ").replace("\r", " ")
            break
        except (UnicodeDecodeError, UnicodeError):
            data = {}
            continue
    if not data:
        data[(1, 1)] = ""
    return _XlsWorkbook({sheetName: _XlsSheetWrapper(sheetName, data)})


def isXlsFile(path: str) -> bool:
    return Path(path).suffix.lower() == ".xls"


def loadWorkbook(path: str) -> _XlsWorkbook:
    """加载工作簿，所有格式均以只读模式打开"""
    suffix = Path(path).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _loadReadOnlyXlsx(path)
    if suffix == ".xls":
        return _loadXls(path)
    if suffix == ".csv":
        return _loadCsv(path)
    raise ValueError(f"Unsupported format: {suffix}")


def getCellValue(ws, row: int, col: int) -> str:
    try:
        val = ws.cell(row=row, column=col).value
        return "" if val is None else str(val)
    except Exception:
        return ""


# ---------- UI ----------

class InputCanceled(Message):
    """用户按 Esc 取消输入"""


class EscapeCancelInput(Input):
    def key_escape(self, event: Key) -> None:
        event.prevent_default()
        self.remove()
        self.post_message(InputCanceled())


class SheetSelectScreen(Screen):
    BINDINGS = [
        Binding("q", "quit", "退出"),
        Binding("enter", "select", "选择"),
    ]

    def __init__(self, workbook, filePath: str, **kwargs):
        super().__init__(**kwargs)
        self.workbook = workbook
        self.filePath = filePath
        self.selectedIndex = 0

    def compose(self) -> ComposeResult:
        yield Header(show_clock=False)
        yield Static(f"[bold]文件:[/] {self.filePath}", id="fileInfo")
        yield Static("[bold]选择 Sheet (方向键移动, Enter 进入):[/]", id="sheetTitle")
        with Vertical(id="sheetList"):
            for i, name in enumerate(self.workbook.sheetnames):
                prefix = "> " if i == self.selectedIndex else "  "
                yield Static(f"{prefix} {name}", id=f"sheet_{i}", classes="sheetItem")
        yield Footer()

    def on_mount(self) -> None:
        self._updateHighlight()

    def _updateHighlight(self) -> None:
        for i, name in enumerate(self.workbook.sheetnames):
            widget = self.query_one(f"#sheet_{i}", Static)
            prefix = "> " if i == self.selectedIndex else "  "
            widget.update(f"{prefix} {name}")

    def key_up(self, event: Key) -> None:
        event.prevent_default()
        if self.selectedIndex > 0:
            self.selectedIndex -= 1
            self._updateHighlight()

    def key_down(self, event: Key) -> None:
        event.prevent_default()
        if self.selectedIndex < len(self.workbook.sheetnames) - 1:
            self.selectedIndex += 1
            self._updateHighlight()

    def on_click(self, event) -> None:
        for i in range(len(self.workbook.sheetnames)):
            try:
                widget = self.query_one(f"#sheet_{i}", Static)
                region = widget.region
                if (
                    region.y <= event.screen_y < region.y + region.height
                    and region.x <= event.screen_x < region.x + region.width
                ):
                    if self.selectedIndex == i:
                        self.action_select()
                    else:
                        self.selectedIndex = i
                        self._updateHighlight()
                    return
            except Exception:
                continue

    def action_select(self) -> None:
        sheetName = self.workbook.sheetnames[self.selectedIndex]
        self.app.push_screen(SheetViewScreen(self.workbook, sheetName, self.filePath))

    def action_quit(self) -> None:
        self.app.exit()


class SheetViewScreen(Screen):
    BINDINGS = [
        Binding("q", "back", "返回"),
        Binding("/", "search", "搜索"),
        Binding("f", "filter", "过滤"),
        Binding("F", "clear_filter", "清除过滤"),
        Binding("g", "goto_row", "跳转行", show=False),
        Binding("ctrl+home", "first_row", "首行", show=False),
        Binding("ctrl+end,G", "last_row", "末行", show=False),
        Binding("home", "first_column", "首列", show=False),
        Binding("end", "last_column", "末列", show=False),
        Binding("n", "next_match", "下一匹配", show=False),
        Binding("p", "prev_match", "上一匹配", show=False),
        Binding("c", "copy_cell", "拷贝"),
        Binding("enter", "enter_row", "进入行"),
    ]

    def __init__(self, workbook, sheetName: str, filePath: str, **kwargs):
        super().__init__(**kwargs)
        self.workbook = workbook
        self.sheetName = sheetName
        self.filePath = filePath
        self.ws = workbook[sheetName]
        self.cursorRow = 1
        self.cursorCol = 1
        self.viewTopRow = 1
        self.viewLeftCol = 1
        self.maxRow = self.ws.max_row or 1
        self.maxCol = self.ws.max_column or 1
        self.isSearchMode = False
        self.isFilterMode = False
        self.isGotoRowMode = False
        self.searchMatches: list[tuple[int, int]] = []
        self.searchMatchIndex = -1
        self.searchQuery: str | None = None
        self.filterQuery: str | None = None
        self.filteredRows: list[int] | None = None
        self._filterViewIndex = 0

    def compose(self) -> ComposeResult:
        yield Header(show_clock=False)
        yield Static("", id="sheetHeader")
        with ScrollableContainer(id="gridContainer", can_focus=False):
            yield Static("", id="gridContent")
        yield Footer()

    def on_mount(self) -> None:
        self._renderGrid()
        self.set_timer(0.05, self._deferredRender)

    def _deferredRender(self) -> None:
        self._renderGrid()

    def on_resize(self, event: Resize) -> None:
        self._renderGrid()

    def _getVisibleRows(self) -> int:
        try:
            container = self.query_one("#gridContainer", ScrollableContainer)
            h = container.size.height if container else 0
            if h <= 0 and self.size.height > 0:
                h = self.size.height - 6
            if self.filteredRows is None and h < 15 and self.size.height > 10:
                h = self.size.height - 6
            return max(1, h - 2)
        except Exception:
            return 20

    def _getVisibleCols(self, rowNumWidth: int = 7) -> int:
        try:
            container = self.query_one("#gridContainer", ScrollableContainer)
            w = container.size.width if container else 0
            if w <= 0 and self.size.width > 0:
                w = self.size.width - 4
            usableWidth = max(0, w - 4 - rowNumWidth - 4)
            return max(1, usableWidth // 15)
        except Exception:
            return 10

    def _formatCell(self, val: str, width: int = 12) -> str:
        displayVal = _formatDisplayValue(val)
        escaped = _escapeForRich(displayVal)
        return _padToDisplayWidth(escaped, width, truncate=True)

    def _getRowsToRender(self) -> list[int]:
        visibleRows = self._getVisibleRows()
        if self.filteredRows is None:
            endRow = (
                min(visibleRows, self.maxRow)
                if self.viewTopRow == 1
                else min(self.viewTopRow + visibleRows - 2, self.maxRow)
            )
            startRow = 2 if self.viewTopRow == 1 else self.viewTopRow
            return list(range(startRow, endRow + 1))
        startIdx = max(0, self._filterViewIndex)
        endIdx = min(startIdx + visibleRows - 1, len(self.filteredRows) - 1)
        return self.filteredRows[startIdx : endIdx + 1]

    def _renderGrid(self) -> None:
        visibleRows = self._getVisibleRows()
        rowNumWidth = max(7, _displayWidth(str(self.maxRow)) + 2)
        visibleCols = self._getVisibleCols(rowNumWidth)
        endCol = min(self.viewLeftCol + visibleCols - 1, self.maxCol)
        cellWidth = 12
        colSep = " │ "

        lines = []
        headerParts = [
            f"[bold white on blue]{_padToDisplayWidth('行号', rowNumWidth, truncate=False)}[/]"
        ]
        for c in range(self.viewLeftCol, endCol + 1):
            val = _formatDisplayValue(getCellValue(self.ws, 1, c))
            cell = self._formatCell(val or get_column_letter(c), cellWidth)
            if self.cursorRow == 1 and c == self.cursorCol:
                headerParts.append(f"{colSep}[bold white on bright_blue]{cell}[/]")
            else:
                headerParts.append(f"{colSep}[bold white on blue]{cell}[/]")
        lines.append("".join(headerParts))

        rowsToRender = self._getRowsToRender()
        for r in rowsToRender:
            rowNumStr = _padToDisplayWidthRight(str(r), rowNumWidth)
            rowParts = [rowNumStr]
            for c in range(self.viewLeftCol, endCol + 1):
                val = getCellValue(self.ws, r, c)
                cell = self._formatCell(val, cellWidth)
                sep = colSep
                if r == self.cursorRow and c == self.cursorCol:
                    rowParts.append(f"{sep}[reverse]{cell}[/reverse]")
                else:
                    rowParts.append(f"{sep}{cell}")
            lines.append("".join(rowParts))

        self.query_one("#gridContent", Static).update("\n".join(lines))

        header = self.query_one("#sheetHeader", Static)
        parts = [
            f"[bold cyan]{Path(self.filePath).name}[/] [dim]→[/] [bold]{self.sheetName}[/]  ",
            f"[dim]行 {self.cursorRow}/{self.maxRow}  列 {self.cursorCol}/{self.maxCol}[/]",
            "  [dim]只读[/]",
        ]
        if self.filteredRows is not None and self.filterQuery:
            parts.append(f"  [yellow]过滤: {self.filterQuery} {len(self.filteredRows)}行[/]")
        if self.searchMatches and self.searchQuery:
            n = self.searchMatchIndex + 1
            m = len(self.searchMatches)
            parts.append(f"  [green]搜索: {self.searchQuery} {n}/{m}[/]")
        rawVal = getCellValue(self.ws, self.cursorRow, self.cursorCol)
        displayVal = _formatDisplayValue(rawVal) if rawVal is not None else ""
        header.update("".join(parts) + f"\n[dim]cellValue:[/] {_escapeForRich(displayVal)}")

    def key_up(self, event: Key) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        event.prevent_default()
        if self.filteredRows is not None:
            try:
                idx = self.filteredRows.index(self.cursorRow)
                if idx > 0:
                    self.cursorRow = self.filteredRows[idx - 1]
                    if idx - 1 < self._filterViewIndex:
                        self._filterViewIndex = max(0, idx - 1)
                    self._renderGrid()
            except ValueError:
                pass
        elif self.cursorRow > 1:
            self.cursorRow -= 1
            if self.cursorRow < self.viewTopRow:
                self.viewTopRow = self.cursorRow
            self._renderGrid()

    def key_down(self, event: Key) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        event.prevent_default()
        if self.filteredRows is not None:
            try:
                idx = self.filteredRows.index(self.cursorRow)
                if idx < len(self.filteredRows) - 1:
                    self.cursorRow = self.filteredRows[idx + 1]
                    visibleRows = self._getVisibleRows()
                    if idx + 1 >= self._filterViewIndex + visibleRows:
                        self._filterViewIndex = idx + 1 - visibleRows + 1
                    self._renderGrid()
            except ValueError:
                if self.filteredRows:
                    self.cursorRow = self.filteredRows[0]
                    self._filterViewIndex = 0
                    self._renderGrid()
        elif self.cursorRow < self.maxRow:
            self.cursorRow += 1
            visibleRows = self._getVisibleRows()
            if self.cursorRow >= self.viewTopRow + visibleRows:
                self.viewTopRow = self.cursorRow - visibleRows + 1
            self._renderGrid()

    def key_left(self, event: Key) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        event.prevent_default()
        if self.cursorCol > 1:
            self.cursorCol -= 1
            if self.cursorCol < self.viewLeftCol:
                self.viewLeftCol = self.cursorCol
            self._renderGrid()

    def key_right(self, event: Key) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        event.prevent_default()
        if self.cursorCol < self.maxCol:
            self.cursorCol += 1
            rowNumWidth = max(7, _displayWidth(str(self.maxRow)) + 2)
            visibleCols = self._getVisibleCols(rowNumWidth)
            if self.cursorCol >= self.viewLeftCol + visibleCols:
                self.viewLeftCol = self.cursorCol - visibleCols + 1
            self._renderGrid()

    def action_back(self) -> None:
        if len(self.app.screen_stack) > 2:
            self.app.pop_screen()
        else:
            self.app.exit(0)

    def action_search(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self.isSearchMode = True
        searchInput = EscapeCancelInput(
            placeholder="搜索: 关键词 或 c:关键词(当前列)  Esc取消",
            id="searchInput",
        )
        self.mount(searchInput)
        searchInput.focus()

    def action_goto_row(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self.isGotoRowMode = True
        gotoInput = EscapeCancelInput(
            placeholder=f"跳转到行号 (1-{self.maxRow})  Esc取消",
            value=str(self.cursorRow),
            id="gotoRowInput",
        )
        self.mount(gotoInput)
        gotoInput.focus()

    def action_filter(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self.isFilterMode = True
        filterInput = EscapeCancelInput(
            placeholder="过滤(当前列): 关键词  Esc取消",
            id="filterInput",
        )
        self.mount(filterInput)
        filterInput.focus()

    def action_clear_filter(self) -> None:
        if self.filteredRows is not None:
            self.filterQuery = None
            self.filteredRows = None
            self._filterViewIndex = 0
            self.viewTopRow = 1
            self._renderGrid()
            self.refresh()
            try:
                self.query_one("#gridContainer", ScrollableContainer).refresh()
            except Exception:
                pass
            self.set_timer(0.1, self._deferredRender)
            self.notify("Filter cleared", timeout=2)

    def action_copy_cell(self) -> None:
        if self.isSearchMode or self.isFilterMode:
            return
        val = getCellValue(self.ws, self.cursorRow, self.cursorCol)
        text = "" if val is None else str(val)
        try:
            self.app.copy_to_clipboard(text)
            self.notify("Copied to clipboard", timeout=2)
        except Exception as e:
            self.notify(f"Copy failed: {e}", severity="error", timeout=3)

    def action_first_row(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self._jumpToRow(1)

    def action_last_row(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self._jumpToRow(self.maxRow)

    def action_first_column(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self.cursorCol = 1
        self.viewLeftCol = 1
        self._renderGrid()

    def action_last_column(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self.cursorCol = self.maxCol
        rowNumWidth = max(7, _displayWidth(str(self.maxRow)) + 2)
        visibleCols = self._getVisibleCols(rowNumWidth)
        self.viewLeftCol = max(1, self.maxCol - visibleCols + 1)
        self._renderGrid()

    def _jumpToRow(self, targetRow: int) -> None:
        targetRow = max(1, min(targetRow, self.maxRow))
        if self.filteredRows is not None:
            idx = bisect.bisect_left(self.filteredRows, targetRow)
            self.cursorRow = (
                self.filteredRows[idx]
                if idx < len(self.filteredRows)
                else self.filteredRows[-1]
            )
            self._filterViewIndex = self.filteredRows.index(self.cursorRow)
        else:
            self.cursorRow = targetRow
            visibleRows = self._getVisibleRows()
            if targetRow < self.viewTopRow or targetRow >= self.viewTopRow + visibleRows:
                self.viewTopRow = max(1, targetRow - visibleRows // 2)
        self._renderGrid()

    def action_enter_row(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self.app.push_screen(
            RowViewScreen(
                self.workbook,
                self.sheetName,
                self.cursorRow,
                self.filePath,
                cursorCol=self.cursorCol,
            )
        )

    def _exitInputMode(self) -> None:
        for inputId in ("searchInput", "filterInput", "gotoRowInput"):
            try:
                self.query_one(f"#{inputId}", Input).remove()
            except Exception:
                pass
        self.isSearchMode = False
        self.isFilterMode = False
        self.isGotoRowMode = False

    def _parseSearchInput(self, raw: str) -> tuple[str, bool]:
        raw = raw.strip()
        lower = raw.lower()
        if lower.startswith("c:") or lower.startswith("c "):
            return raw[2:].strip(), True
        if lower.startswith("列:") or lower.startswith("列 "):
            return raw[2:].strip(), True
        if lower.startswith("col:") or lower.startswith("col "):
            return raw[4:].strip(), True
        return raw, False

    def on_input_submitted(self, event: Input.Submitted) -> None:
        if event.input.id == "searchInput":
            raw = event.value.strip()
            self._exitInputMode()
            if raw:
                self._doSearch(raw)
        elif event.input.id == "filterInput":
            raw = event.value.strip()
            self._exitInputMode()
            if raw:
                self._doFilter(raw)
            else:
                self.action_clear_filter()
        elif event.input.id == "gotoRowInput":
            raw = event.value.strip().lower()
            self._exitInputMode()
            if raw:
                if raw == "g":
                    targetRow = 1
                elif raw == "e":
                    targetRow = self.maxRow
                else:
                    try:
                        targetRow = max(1, min(int(raw), self.maxRow))
                    except ValueError:
                        self.notify("Invalid row number", severity="error", timeout=2)
                        return
                self._jumpToRow(targetRow)
                self.notify(f"Jumped to row {self.cursorRow}", timeout=2)

    def _doFilter(self, raw: str) -> None:
        query = raw.strip()
        if not query:
            self.action_clear_filter()
            return
        queryLower = query.lower()
        c = self.cursorCol
        matchingRows: set[int] = set()
        for r in range(1, self.maxRow + 1):
            val = getCellValue(self.ws, r, c)
            if val is not None and queryLower in str(val).lower():
                matchingRows.add(r)
        self.filterQuery = query
        self.filteredRows = sorted(matchingRows) if matchingRows else []
        self._filterViewIndex = 0
        if not self.filteredRows:
            self.notify("No rows match filter", timeout=2)
            return
        self.cursorRow = self.filteredRows[0]
        self._renderGrid()
        self.notify(f"Filter: {len(self.filteredRows)} rows", timeout=2)

    def _doSearch(self, raw: str) -> None:
        query, columnOnly = self._parseSearchInput(raw)
        if not query:
            self.notify("Empty search query", timeout=2)
            return
        queryLower = query.lower()
        matches: list[tuple[int, int]] = []
        if columnOnly:
            c = self.cursorCol
            for r in range(1, self.maxRow + 1):
                val = getCellValue(self.ws, r, c)
                if queryLower in val.lower():
                    matches.append((r, c))
            scope = "column"
        else:
            for r in range(1, self.maxRow + 1):
                for c in range(1, self.maxCol + 1):
                    val = getCellValue(self.ws, r, c)
                    if queryLower in val.lower():
                        matches.append((r, c))
            scope = "global"
        self.searchMatches = matches
        self.searchMatchIndex = -1
        self.searchQuery = raw.strip()
        if not matches:
            self.notify(f"No match in {scope} search", timeout=2)
            return
        self._goToMatchByIndex(0)
        self.notify(f"{len(matches)} matches", timeout=2)

    def action_next_match(self) -> None:
        if not self.searchMatches:
            self.notify("No search results, press / to search", timeout=2)
            return
        self._goToMatchByIndex((self.searchMatchIndex + 1) % len(self.searchMatches))

    def action_prev_match(self) -> None:
        if not self.searchMatches:
            self.notify("No search results, press / to search", timeout=2)
            return
        self._goToMatchByIndex((self.searchMatchIndex - 1) % len(self.searchMatches))

    def _goToMatchByIndex(self, idx: int) -> None:
        if not self.searchMatches:
            return
        self.searchMatchIndex = idx % len(self.searchMatches)
        r, c = self.searchMatches[self.searchMatchIndex]
        self.cursorRow, self.cursorCol = r, c
        self.viewTopRow = max(1, r - 2)
        self.viewLeftCol = max(1, c - 2)
        self._renderGrid()

    def on_input_canceled(self, event: InputCanceled) -> None:
        self.isSearchMode = False
        self.isFilterMode = False
        self.isGotoRowMode = False
        self._renderGrid()

    def key_escape(self, event: Key) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            event.prevent_default()
            self._exitInputMode()
            self._renderGrid()

    def on_mouse_scroll_down(self, event) -> None:
        event.stop()
        if self.filteredRows is not None:
            try:
                idx = self.filteredRows.index(self.cursorRow)
                if idx < len(self.filteredRows) - 1:
                    self.cursorRow = self.filteredRows[idx + 1]
                    visibleRows = self._getVisibleRows()
                    if idx + 1 >= self._filterViewIndex + visibleRows:
                        self._filterViewIndex = idx + 1 - visibleRows + 1
                    self._renderGrid()
            except ValueError:
                pass
        elif self.cursorRow < self.maxRow:
            self.cursorRow += 1
            visibleRows = self._getVisibleRows()
            if self.cursorRow >= self.viewTopRow + visibleRows:
                self.viewTopRow = self.cursorRow - visibleRows + 1
            self._renderGrid()

    def on_mouse_scroll_up(self, event) -> None:
        event.stop()
        if self.filteredRows is not None:
            try:
                idx = self.filteredRows.index(self.cursorRow)
                if idx > 0:
                    self.cursorRow = self.filteredRows[idx - 1]
                    if idx - 1 < self._filterViewIndex:
                        self._filterViewIndex = max(0, idx - 1)
                    self._renderGrid()
            except ValueError:
                pass
        elif self.cursorRow > 1:
            self.cursorRow -= 1
            if self.cursorRow < self.viewTopRow:
                self.viewTopRow = self.cursorRow
            self._renderGrid()

    def on_click(self, event) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        try:
            gridWidget = self.query_one("#gridContent", Static)
            contentRegion = gridWidget.content_region
        except Exception:
            return
        contentX = event.screen_x - contentRegion.x
        contentY = event.screen_y - contentRegion.y
        if contentX < 0 or contentY < 0:
            return
        if contentX >= contentRegion.width or contentY >= contentRegion.height:
            return
        self._clickToCell(contentX, contentY)

    def _clickToCell(self, contentX: int, contentY: int) -> None:
        """根据 #gridContent 内容区坐标映射到 (row, col) 并移动光标"""
        rowNumWidth = max(7, _displayWidth(str(self.maxRow)) + 2)
        cellWidth = 12
        colSepWidth = 3  # " │ " 占 3 个终端列

        # y=0 → 表头行(row 1)，y>=1 → 对应渲染数据行
        if contentY == 0:
            clickedRow = 1
        else:
            rowsToRender = self._getRowsToRender()
            dataIdx = contentY - 1
            if dataIdx >= len(rowsToRender):
                return
            clickedRow = rowsToRender[dataIdx]

        # x < rowNumWidth → 点在行号区，只跳行不动列
        if contentX < rowNumWidth:
            self.cursorRow = clickedRow
            self._renderGrid()
            return

        colOffset = (contentX - rowNumWidth) // (colSepWidth + cellWidth)
        clickedCol = max(1, min(self.viewLeftCol + colOffset, self.maxCol))

        self.cursorRow = clickedRow
        self.cursorCol = clickedCol
        self._renderGrid()


class RowViewScreen(Screen):
    BINDINGS = [
        Binding("q", "back", "返回"),
        Binding("c", "copy_cell", "拷贝"),
    ]

    def __init__(
        self,
        workbook,
        sheetName: str,
        rowIndex: int,
        filePath: str,
        cursorCol: int = 1,
        **kwargs,
    ):
        super().__init__(**kwargs)
        self.workbook = workbook
        self.sheetName = sheetName
        self.rowIndex = rowIndex
        self.filePath = filePath
        self.ws = workbook[sheetName]
        self.maxCol = self.ws.max_column or 1
        self._initCursorCol = min(max(1, cursorCol), self.maxCol)

    def compose(self) -> ComposeResult:
        yield Header(show_clock=False)
        yield Static(
            f"[bold green]{Path(self.filePath).name}[/] [dim]→[/] [bold]{self.sheetName}[/] "
            f"[dim]→[/] [bold]行{self.rowIndex}[/]  [dim]↑↓选择 c拷贝 q返回[/]",
            id="rowBanner",
        )
        yield DataTable(id="rowTable", cursor_type="row", zebra_stripes=True)
        yield Footer()

    def on_mount(self) -> None:
        rows: list[tuple[str, str]] = []
        for c in range(1, self.maxCol + 1):
            fieldName = getCellValue(self.ws, 1, c) or get_column_letter(c)
            val = _formatDisplayValue(getCellValue(self.ws, self.rowIndex, c))
            rows.append((fieldName, val))

        nameWidth = max((_displayWidth(r[0]) for r in rows), default=4)
        nameWidth = max(nameWidth, _displayWidth("字段名"))
        valWidth = max((_displayWidth(r[1]) for r in rows), default=4)
        valWidth = max(valWidth, _displayWidth("值"))

        table = self.query_one(DataTable)
        table.add_column("字段名", width=nameWidth)
        table.add_column("值", width=valWidth)
        for fieldName, val in rows:
            table.add_row(fieldName, val)
        table.move_cursor(row=self._initCursorCol - 1)
        table.focus()

    @property
    def cursorCol(self) -> int:
        try:
            return self.query_one(DataTable).cursor_row + 1
        except Exception:
            return 1

    def action_back(self) -> None:
        self.app.pop_screen()

    def action_copy_cell(self) -> None:
        val = getCellValue(self.ws, self.rowIndex, self.cursorCol)
        text = "" if val is None else str(val)
        try:
            self.app.copy_to_clipboard(text)
            self.notify("Copied to clipboard", timeout=2)
        except Exception as e:
            self.notify(f"Copy failed: {e}", severity="error", timeout=3)

    def key_c(self, event: Key) -> None:
        event.prevent_default()
        event.stop()
        self.action_copy_cell()

    def key_q(self, event: Key) -> None:
        event.prevent_default()
        event.stop()
        self.app.pop_screen()

    def key_escape(self, event: Key) -> None:
        event.prevent_default()
        event.stop()
        self.app.pop_screen()


class ExcelTuiApp(App):
    TITLE = "Excel TUI"
    CSS = """
    #fileInfo { padding: 1; margin: 1; }
    #sheetTitle { padding: 1; margin: 1; }
    .sheetItem { padding: 0 1; margin: 0 1; }
    #gridContainer {
        height: 1fr;
        min-height: 20;
        padding: 1 2; border: solid $primary;
        overflow-x: auto; overflow-y: auto;
    }
    #gridContent { padding: 1; }
    #rowTable { height: 1fr; }
    #sheetHeader, #rowBanner { padding: 1; }
    #searchInput, #filterInput, #gotoRowInput {
        dock: top; margin: 1; width: 50;
    }
    """

    def __init__(self, filePath: str, **kwargs):
        super().__init__(**kwargs)
        self.filePath = filePath
        self.workbook = None

    def on_mount(self) -> None:
        try:
            self.workbook = loadWorkbook(self.filePath)
        except Exception as e:
            self.notify(f"Failed to load file: {e}", severity="error")
            self.exit(1)
            return
        if len(self.workbook.sheetnames) == 1:
            sheetName = self.workbook.sheetnames[0]
            self.push_screen(SheetViewScreen(self.workbook, sheetName, self.filePath))
        else:
            self.push_screen(SheetSelectScreen(self.workbook, self.filePath))


def main() -> int:
    parser = argparse.ArgumentParser(description="Excel/CSV TUI - 终端界面只读查看表格文件")
    parser.add_argument("file", type=str, help="文件路径 (.xlsx/.xlsm/.xls/.csv)")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(f"Error: file not found: {path}", file=sys.stderr)
        return 1

    supported = {".xlsx", ".xlsm", ".xls", ".csv"}
    if path.suffix.lower() not in supported:
        print(f"Error: unsupported format '{path.suffix}', supported: {', '.join(sorted(supported))}", file=sys.stderr)
        return 1

    app = ExcelTuiApp(str(path.resolve()))
    result = app.run()
    return result if isinstance(result, int) else 0


if __name__ == "__main__":
    sys.exit(main())
