#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TUI 界面：Sheet 选择、表格浏览、行详情
"""

import bisect
from pathlib import Path

from openpyxl.utils import get_column_letter
from textual.app import App, ComposeResult
from textual.binding import Binding
from textual.containers import Vertical, ScrollableContainer
from textual.events import Key, Resize
from textual.message import Message
from textual.screen import Screen
from textual.widgets import DataTable, Footer, Header, Input, Static

from .utils import (
    displayWidth,
    padToDisplayWidth,
    padToDisplayWidthRight,
    formatDisplayValue,
    escapeForRich,
    loadColWidthsConfig,
    saveColWidthsConfig,
)
from .workbook import getCellValue


# ---------- 通用输入组件 ----------

class InputCanceled(Message):
    """用户按 Esc 取消输入"""


class EscapeCancelInput(Input):
    def key_escape(self, event: Key) -> None:
        event.prevent_default()
        self.remove()
        self.post_message(InputCanceled())


# ---------- Sheet 选择界面 ----------

class SheetSelectScreen(Screen):
    BINDINGS = [
        Binding("q", "quit", "退出"),
        Binding("enter", "select", "选择"),
    ]

    def __init__(self, workbook, filePath: str, schemaData: dict | None = None, **kwargs):
        super().__init__(**kwargs)
        self.workbook = workbook
        self.filePath = filePath
        self.schemaData = schemaData
        self.selectedIndex = 0

    def _sheetLabel(self, name: str, selected: bool) -> str:
        prefix = "> " if selected else "  "
        entry = (self.schemaData or {}).get(name)
        suffix = f"  [dim]→ {entry['exportName']}[/]" if entry else ""
        return f"{prefix} {name}{suffix}"

    def compose(self) -> ComposeResult:
        yield Header(show_clock=False)
        yield Static(f"[bold]文件:[/] {self.filePath}", id="fileInfo")
        yield Static("[bold]选择 Sheet (方向键移动, Enter 进入):[/]", id="sheetTitle")
        with Vertical(id="sheetList"):
            for i, name in enumerate(self.workbook.sheetnames):
                yield Static(
                    self._sheetLabel(name, i == self.selectedIndex),
                    id=f"sheet_{i}",
                    classes="sheetItem",
                )
        yield Footer()

    def on_mount(self) -> None:
        self._updateHighlight()

    def _updateHighlight(self) -> None:
        for i, name in enumerate(self.workbook.sheetnames):
            widget = self.query_one(f"#sheet_{i}", Static)
            widget.update(self._sheetLabel(name, i == self.selectedIndex))

    def key_up(self, event: Key) -> None:
        event.prevent_default()
        if self.selectedIndex > 0:
            self.selectedIndex -= 1
            self._updateHighlight()

    def key_down(self, event: Key) -> None:
        event.prevent_default()
        if self.selectedIndex < len(self.workbook.sheetnames) - 1:
            self.selectedIndex += 1
            self._updateHighlight()

    def on_click(self, event) -> None:
        for i in range(len(self.workbook.sheetnames)):
            try:
                widget = self.query_one(f"#sheet_{i}", Static)
                region = widget.region
                if (
                    region.y <= event.screen_y < region.y + region.height
                    and region.x <= event.screen_x < region.x + region.width
                ):
                    if self.selectedIndex == i:
                        self.action_select()
                    else:
                        self.selectedIndex = i
                        self._updateHighlight()
                    return
            except Exception:
                continue

    def action_select(self) -> None:
        sheetName = self.workbook.sheetnames[self.selectedIndex]
        self.app.push_screen(
            SheetViewScreen(self.workbook, sheetName, self.filePath, schemaData=self.schemaData)
        )

    def action_quit(self) -> None:
        self.app.exit()


# ---------- 表格浏览界面 ----------

class SheetViewScreen(Screen):
    BINDINGS = [
        Binding("q", "back", "返回"),
        Binding("/", "search", "搜索"),
        Binding("f", "filter", "过滤"),
        Binding("F", "clear_filter", "清除过滤"),
        Binding("s", "toggle_schema", "Schema"),
        Binding("g", "goto_row", "跳转行", show=False),
        Binding("ctrl+home", "first_row", "首行", show=False),
        Binding("ctrl+end,G", "last_row", "末行", show=False),
        Binding("home", "first_column", "首列", show=False),
        Binding("end", "last_column", "末列", show=False),
        Binding("n", "next_match", "下一匹配", show=False),
        Binding("p", "prev_match", "上一匹配", show=False),
        Binding("c", "copy_cell", "拷贝"),
        Binding("enter", "enter_row", "进入行"),
        Binding("[", "decrease_col_width", "缩窄列"),
        Binding("]", "increase_col_width", "加宽列"),
    ]

    def __init__(
        self,
        workbook,
        sheetName: str,
        filePath: str,
        schemaData: dict | None = None,
        **kwargs,
    ):
        super().__init__(**kwargs)
        self.workbook = workbook
        self.sheetName = sheetName
        self.filePath = filePath
        self.schemaData = schemaData
        self.ws = workbook[sheetName]
        self.cursorRow = 1
        self.cursorCol = 1
        self.viewTopRow = 1
        self.viewLeftCol = 1
        self.maxRow = self.ws.max_row or 1
        self.maxCol = self.ws.max_column or 1
        self.isSearchMode = False
        self.isFilterMode = False
        self.isGotoRowMode = False
        self.searchMatches: list[tuple[int, int]] = []
        self.searchMatchIndex = -1
        self.searchQuery: str | None = None
        self.filterQuery: str | None = None
        self.filteredRows: list[int] | None = None
        self._filterViewIndex = 0
        self.columnWidths: dict[int, int] = {}
        self._defaultCellWidth = 12
        self.schemaVisible: bool = bool(schemaData and sheetName in schemaData)

    def compose(self) -> ComposeResult:
        yield Header(show_clock=False)
        yield Static("", id="sheetHeader")
        with Vertical(id="schemaPanel"):
            yield Static("", id="schemaEnRow")
            yield Static("", id="schemaTypeRow")
        with ScrollableContainer(id="gridContainer", can_focus=False):
            yield Static("", id="gridContent")
        yield Footer()

    def on_mount(self) -> None:
        self._loadColumnWidths()
        try:
            self.query_one("#schemaPanel").display = self.schemaVisible
        except Exception:
            pass
        self._renderGrid()
        self.set_timer(0.05, self._deferredRender)

    def _deferredRender(self) -> None:
        self._renderGrid()

    def on_resize(self, event: Resize) -> None:
        self._renderGrid()

    # ---------- 布局计算 ----------

    def _getVisibleRows(self) -> int:
        try:
            container = self.query_one("#gridContainer", ScrollableContainer)
            h = container.size.height if container else 0
            if h <= 0 and self.size.height > 0:
                h = self.size.height - 6
            if self.filteredRows is None and h < 15 and self.size.height > 10:
                h = self.size.height - 6
            return max(1, h - 2)
        except Exception:
            return 20

    def _getColWidth(self, col: int) -> int:
        return self.columnWidths.get(col, self._defaultCellWidth)

    def _colConfigKey(self, col: int) -> str:
        name = getCellValue(self.ws, 1, col)
        return name.strip() if name and name.strip() else f"col_{col}"

    def _fileConfigKey(self) -> str:
        return str(Path(self.filePath).resolve())

    def _loadColumnWidths(self) -> None:
        data = loadColWidthsConfig()
        sheetData = data.get(self._fileConfigKey(), {}).get(self.sheetName, {})
        for col in range(1, self.maxCol + 1):
            key = self._colConfigKey(col)
            if key in sheetData:
                w = sheetData[key]
                if isinstance(w, int) and 4 <= w <= 80:
                    self.columnWidths[col] = w

    def _saveColumnWidths(self) -> None:
        data = loadColWidthsConfig()
        fileKey = self._fileConfigKey()
        if fileKey not in data:
            data[fileKey] = {}
        sheetWidths: dict[str, int] = {}
        for col, width in self.columnWidths.items():
            sheetWidths[self._colConfigKey(col)] = width
        data[fileKey][self.sheetName] = sheetWidths
        saveColWidthsConfig(data)

    def _computeEndCol(self, rowNumWidth: int) -> int:
        """从 viewLeftCol 开始，按各列实际宽度计算最后一个可见列"""
        try:
            container = self.query_one("#gridContainer", ScrollableContainer)
            w = container.size.width if container else 0
            if w <= 0 and self.size.width > 0:
                w = self.size.width - 4
            usableWidth = max(1, w - 4 - rowNumWidth - 4)
        except Exception:
            usableWidth = 80
        colSepWidth = 3
        used = 0
        for c in range(self.viewLeftCol, self.maxCol + 1):
            if c > self.viewLeftCol:
                used += colSepWidth
            used += self._getColWidth(c)
            if used > usableWidth and c > self.viewLeftCol:
                return c - 1
        return self.maxCol

    def _computeViewLeftToShowCol(self, targetCol: int, rowNumWidth: int) -> int:
        """计算 viewLeftCol，使 targetCol 刚好出现在视口右侧"""
        try:
            container = self.query_one("#gridContainer", ScrollableContainer)
            w = container.size.width if container else 0
            if w <= 0 and self.size.width > 0:
                w = self.size.width - 4
            usableWidth = max(1, w - 4 - rowNumWidth - 4)
        except Exception:
            usableWidth = 80
        colSepWidth = 3
        used = self._getColWidth(targetCol)
        for c in range(targetCol - 1, 0, -1):
            used += colSepWidth + self._getColWidth(c)
            if used > usableWidth:
                return c + 1
        return 1

    def _formatCell(self, val: str, width: int = 12) -> str:
        displayVal = formatDisplayValue(val)
        escaped = escapeForRich(displayVal)
        return padToDisplayWidth(escaped, width, truncate=True)

    def _getRowsToRender(self) -> list[int]:
        visibleRows = self._getVisibleRows()
        if self.filteredRows is None:
            endRow = (
                min(visibleRows, self.maxRow)
                if self.viewTopRow == 1
                else min(self.viewTopRow + visibleRows - 2, self.maxRow)
            )
            startRow = 2 if self.viewTopRow == 1 else self.viewTopRow
            return list(range(startRow, endRow + 1))
        startIdx = max(0, self._filterViewIndex)
        endIdx = min(startIdx + visibleRows - 1, len(self.filteredRows) - 1)
        return self.filteredRows[startIdx : endIdx + 1]

    # ---------- 渲染 ----------

    def _renderGrid(self) -> None:
        rowNumWidth = max(7, displayWidth(str(self.maxRow)) + 2)
        endCol = self._computeEndCol(rowNumWidth)
        colSep = " │ "

        lines = []
        headerParts = [
            f"[bold white on blue]{padToDisplayWidth('行号', rowNumWidth, truncate=False)}[/]"
        ]
        for c in range(self.viewLeftCol, endCol + 1):
            colW = self._getColWidth(c)
            val = formatDisplayValue(getCellValue(self.ws, 1, c))
            cell = self._formatCell(val or get_column_letter(c), colW)
            if self.cursorRow == 1 and c == self.cursorCol:
                headerParts.append(f"{colSep}[bold white on bright_blue]{cell}[/]")
            else:
                headerParts.append(f"{colSep}[bold white on blue]{cell}[/]")
        lines.append("".join(headerParts))

        rowsToRender = self._getRowsToRender()
        for r in rowsToRender:
            rowNumStr = padToDisplayWidthRight(str(r), rowNumWidth)
            rowParts = [rowNumStr]
            for c in range(self.viewLeftCol, endCol + 1):
                colW = self._getColWidth(c)
                val = getCellValue(self.ws, r, c)
                cell = self._formatCell(val, colW)
                if r == self.cursorRow and c == self.cursorCol:
                    rowParts.append(f"{colSep}[reverse]{cell}[/reverse]")
                else:
                    rowParts.append(f"{colSep}{cell}")
            lines.append("".join(rowParts))

        self.query_one("#gridContent", Static).update("\n".join(lines))

        header = self.query_one("#sheetHeader", Static)
        parts = [
            f"[bold cyan]{Path(self.filePath).name}[/] [dim]→[/] [bold]{self.sheetName}[/]  ",
            f"[dim]行 {self.cursorRow}/{self.maxRow}  列 {self.cursorCol}/{self.maxCol}[/]",
            "  [dim]只读[/]",
        ]
        if self.filteredRows is not None and self.filterQuery:
            parts.append(f"  [yellow]过滤: {self.filterQuery} {len(self.filteredRows)}行[/]")
        if self.searchMatches and self.searchQuery:
            n = self.searchMatchIndex + 1
            m = len(self.searchMatches)
            parts.append(f"  [green]搜索: {self.searchQuery} {n}/{m}[/]")
        rawVal = getCellValue(self.ws, self.cursorRow, self.cursorCol)
        displayVal = formatDisplayValue(rawVal) if rawVal is not None else ""
        header.update("".join(parts) + f"\n[dim]cellValue:[/] {escapeForRich(displayVal)}")
        self._renderSchemaRows()

    def _renderSchemaRows(self) -> None:
        """渲染 schema 英文字段名行和类型行（固定在网格上方，不随数据滚动）"""
        if not self.schemaVisible:
            return
        try:
            enWidget = self.query_one("#schemaEnRow", Static)
            typeWidget = self.query_one("#schemaTypeRow", Static)
        except Exception:
            return

        schemaEntry = (self.schemaData or {}).get(self.sheetName)
        rowNumWidth = max(7, displayWidth(str(self.maxRow)) + 2)
        endCol = self._computeEndCol(rowNumWidth)
        colSep = " │ "

        enParts: list[str] = [" " * rowNumWidth]
        typeParts: list[str] = [" " * rowNumWidth]
        for c in range(self.viewLeftCol, endCol + 1):
            colW = self._getColWidth(c)
            enName = ""
            typeName = ""
            if schemaEntry:
                cnName = formatDisplayValue(getCellValue(self.ws, 1, c))
                mapping = schemaEntry["cnToEn"].get(cnName)
                if mapping:
                    enName, typeName = mapping
            enParts.append(f"{colSep}[bold cyan]{self._formatCell(enName, colW)}[/]")
            typeParts.append(f"{colSep}[yellow]{self._formatCell(typeName, colW)}[/]")

        enWidget.update("".join(enParts))
        typeWidget.update("".join(typeParts))

    def action_toggle_schema(self) -> None:
        if not (self.schemaData and self.sheetName in self.schemaData):
            self.notify("当前 Sheet 无 schema 配置", timeout=2)
            return
        self.schemaVisible = not self.schemaVisible
        try:
            self.query_one("#schemaPanel").display = self.schemaVisible
        except Exception:
            pass

    # ---------- 键盘导航 ----------

    def key_up(self, event: Key) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        event.prevent_default()
        if self.filteredRows is not None:
            try:
                idx = self.filteredRows.index(self.cursorRow)
                if idx > 0:
                    self.cursorRow = self.filteredRows[idx - 1]
                    if idx - 1 < self._filterViewIndex:
                        self._filterViewIndex = max(0, idx - 1)
                    self._renderGrid()
            except ValueError:
                pass
        elif self.cursorRow > 1:
            self.cursorRow -= 1
            if self.cursorRow < self.viewTopRow:
                self.viewTopRow = self.cursorRow
            self._renderGrid()

    def key_down(self, event: Key) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        event.prevent_default()
        if self.filteredRows is not None:
            try:
                idx = self.filteredRows.index(self.cursorRow)
                if idx < len(self.filteredRows) - 1:
                    self.cursorRow = self.filteredRows[idx + 1]
                    visibleRows = self._getVisibleRows()
                    if idx + 1 >= self._filterViewIndex + visibleRows:
                        self._filterViewIndex = idx + 1 - visibleRows + 1
                    self._renderGrid()
            except ValueError:
                if self.filteredRows:
                    self.cursorRow = self.filteredRows[0]
                    self._filterViewIndex = 0
                    self._renderGrid()
        elif self.cursorRow < self.maxRow:
            self.cursorRow += 1
            visibleRows = self._getVisibleRows()
            if self.cursorRow >= self.viewTopRow + visibleRows:
                self.viewTopRow = self.cursorRow - visibleRows + 1
            self._renderGrid()

    def key_left(self, event: Key) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        event.prevent_default()
        if self.cursorCol > 1:
            self.cursorCol -= 1
            if self.cursorCol < self.viewLeftCol:
                self.viewLeftCol = self.cursorCol
            self._renderGrid()

    def key_right(self, event: Key) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        event.prevent_default()
        if self.cursorCol < self.maxCol:
            self.cursorCol += 1
            rowNumWidth = max(7, displayWidth(str(self.maxRow)) + 2)
            endCol = self._computeEndCol(rowNumWidth)
            if self.cursorCol > endCol:
                self.viewLeftCol = self._computeViewLeftToShowCol(self.cursorCol, rowNumWidth)
            self._renderGrid()

    # ---------- 动作 ----------

    def action_back(self) -> None:
        if len(self.app.screen_stack) > 2:
            self.app.pop_screen()
        else:
            self.app.exit(0)

    def action_search(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self.isSearchMode = True
        searchInput = EscapeCancelInput(
            placeholder="搜索: 关键词 或 c:关键词(当前列)  Esc取消",
            id="searchInput",
        )
        self.mount(searchInput)
        searchInput.focus()

    def action_goto_row(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self.isGotoRowMode = True
        gotoInput = EscapeCancelInput(
            placeholder=f"跳转到行号 (1-{self.maxRow})  Esc取消",
            value=str(self.cursorRow),
            id="gotoRowInput",
        )
        self.mount(gotoInput)
        gotoInput.focus()

    def action_filter(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self.isFilterMode = True
        filterInput = EscapeCancelInput(
            placeholder="过滤(当前列): 关键词  Esc取消",
            id="filterInput",
        )
        self.mount(filterInput)
        filterInput.focus()

    def action_clear_filter(self) -> None:
        if self.filteredRows is not None:
            self.filterQuery = None
            self.filteredRows = None
            self._filterViewIndex = 0
            self.viewTopRow = 1
            self._renderGrid()
            self.refresh()
            try:
                self.query_one("#gridContainer", ScrollableContainer).refresh()
            except Exception:
                pass
            self.set_timer(0.1, self._deferredRender)
            self.notify("Filter cleared", timeout=2)

    def action_copy_cell(self) -> None:
        if self.isSearchMode or self.isFilterMode:
            return
        val = getCellValue(self.ws, self.cursorRow, self.cursorCol)
        text = "" if val is None else str(val)
        try:
            self.app.copy_to_clipboard(text)
            self.notify("Copied to clipboard", timeout=2)
        except Exception as e:
            self.notify(f"Copy failed: {e}", severity="error", timeout=3)

    def action_first_row(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self._jumpToRow(1)

    def action_last_row(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self._jumpToRow(self.maxRow)

    def action_first_column(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self.cursorCol = 1
        self.viewLeftCol = 1
        self._renderGrid()

    def action_last_column(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self.cursorCol = self.maxCol
        rowNumWidth = max(7, displayWidth(str(self.maxRow)) + 2)
        self.viewLeftCol = self._computeViewLeftToShowCol(self.maxCol, rowNumWidth)
        self._renderGrid()

    def _jumpToRow(self, targetRow: int) -> None:
        targetRow = max(1, min(targetRow, self.maxRow))
        if self.filteredRows is not None:
            idx = bisect.bisect_left(self.filteredRows, targetRow)
            self.cursorRow = (
                self.filteredRows[idx]
                if idx < len(self.filteredRows)
                else self.filteredRows[-1]
            )
            self._filterViewIndex = self.filteredRows.index(self.cursorRow)
        else:
            self.cursorRow = targetRow
            visibleRows = self._getVisibleRows()
            if targetRow < self.viewTopRow or targetRow >= self.viewTopRow + visibleRows:
                self.viewTopRow = max(1, targetRow - visibleRows // 2)
        self._renderGrid()

    def action_enter_row(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        self.app.push_screen(
            RowViewScreen(
                self.workbook,
                self.sheetName,
                self.cursorRow,
                self.filePath,
                cursorCol=self.cursorCol,
            )
        )

    def action_decrease_col_width(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        col = self.cursorCol
        current = self._getColWidth(col)
        self.columnWidths[col] = max(4, current - 2)
        self._renderGrid()
        self._saveColumnWidths()
        self.notify(f"列 {col} 宽度: {self.columnWidths[col]}", timeout=1)

    def action_increase_col_width(self) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        col = self.cursorCol
        current = self._getColWidth(col)
        self.columnWidths[col] = min(80, current + 2)
        self._renderGrid()
        self._saveColumnWidths()
        self.notify(f"列 {col} 宽度: {self.columnWidths[col]}", timeout=1)

    # ---------- 输入模式 ----------

    def _exitInputMode(self) -> None:
        for inputId in ("searchInput", "filterInput", "gotoRowInput"):
            try:
                self.query_one(f"#{inputId}", Input).remove()
            except Exception:
                pass
        self.isSearchMode = False
        self.isFilterMode = False
        self.isGotoRowMode = False

    def _parseSearchInput(self, raw: str) -> tuple[str, bool]:
        raw = raw.strip()
        lower = raw.lower()
        if lower.startswith("c:") or lower.startswith("c "):
            return raw[2:].strip(), True
        if lower.startswith("列:") or lower.startswith("列 "):
            return raw[2:].strip(), True
        if lower.startswith("col:") or lower.startswith("col "):
            return raw[4:].strip(), True
        return raw, False

    def on_input_submitted(self, event: Input.Submitted) -> None:
        if event.input.id == "searchInput":
            raw = event.value.strip()
            self._exitInputMode()
            if raw:
                self._doSearch(raw)
        elif event.input.id == "filterInput":
            raw = event.value.strip()
            self._exitInputMode()
            if raw:
                self._doFilter(raw)
            else:
                self.action_clear_filter()
        elif event.input.id == "gotoRowInput":
            raw = event.value.strip().lower()
            self._exitInputMode()
            if raw:
                if raw == "g":
                    targetRow = 1
                elif raw == "e":
                    targetRow = self.maxRow
                else:
                    try:
                        targetRow = max(1, min(int(raw), self.maxRow))
                    except ValueError:
                        self.notify("Invalid row number", severity="error", timeout=2)
                        return
                self._jumpToRow(targetRow)
                self.notify(f"Jumped to row {self.cursorRow}", timeout=2)

    def _doFilter(self, raw: str) -> None:
        query = raw.strip()
        if not query:
            self.action_clear_filter()
            return
        queryLower = query.lower()
        c = self.cursorCol
        matchingRows: set[int] = set()
        for r in range(1, self.maxRow + 1):
            val = getCellValue(self.ws, r, c)
            if val is not None and queryLower in str(val).lower():
                matchingRows.add(r)
        self.filterQuery = query
        self.filteredRows = sorted(matchingRows) if matchingRows else []
        self._filterViewIndex = 0
        if not self.filteredRows:
            self.notify("No rows match filter", timeout=2)
            return
        self.cursorRow = self.filteredRows[0]
        self._renderGrid()
        self.notify(f"Filter: {len(self.filteredRows)} rows", timeout=2)

    def _doSearch(self, raw: str) -> None:
        query, columnOnly = self._parseSearchInput(raw)
        if not query:
            self.notify("Empty search query", timeout=2)
            return
        queryLower = query.lower()
        matches: list[tuple[int, int]] = []
        if columnOnly:
            c = self.cursorCol
            for r in range(1, self.maxRow + 1):
                val = getCellValue(self.ws, r, c)
                if queryLower in val.lower():
                    matches.append((r, c))
            scope = "column"
        else:
            for r in range(1, self.maxRow + 1):
                for c in range(1, self.maxCol + 1):
                    val = getCellValue(self.ws, r, c)
                    if queryLower in val.lower():
                        matches.append((r, c))
            scope = "global"
        self.searchMatches = matches
        self.searchMatchIndex = -1
        self.searchQuery = raw.strip()
        if not matches:
            self.notify(f"No match in {scope} search", timeout=2)
            return
        self._goToMatchByIndex(0)
        self.notify(f"{len(matches)} matches", timeout=2)

    def action_next_match(self) -> None:
        if not self.searchMatches:
            self.notify("No search results, press / to search", timeout=2)
            return
        self._goToMatchByIndex((self.searchMatchIndex + 1) % len(self.searchMatches))

    def action_prev_match(self) -> None:
        if not self.searchMatches:
            self.notify("No search results, press / to search", timeout=2)
            return
        self._goToMatchByIndex((self.searchMatchIndex - 1) % len(self.searchMatches))

    def _goToMatchByIndex(self, idx: int) -> None:
        if not self.searchMatches:
            return
        self.searchMatchIndex = idx % len(self.searchMatches)
        r, c = self.searchMatches[self.searchMatchIndex]
        self.cursorRow, self.cursorCol = r, c
        self.viewTopRow = max(1, r - 2)
        self.viewLeftCol = max(1, c - 2)
        self._renderGrid()

    def on_input_canceled(self, event: InputCanceled) -> None:
        self.isSearchMode = False
        self.isFilterMode = False
        self.isGotoRowMode = False
        self._renderGrid()

    def key_escape(self, event: Key) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            event.prevent_default()
            self._exitInputMode()
            self._renderGrid()

    # ---------- 鼠标 ----------

    def on_mouse_scroll_down(self, event) -> None:
        event.stop()
        if self.filteredRows is not None:
            try:
                idx = self.filteredRows.index(self.cursorRow)
                if idx < len(self.filteredRows) - 1:
                    self.cursorRow = self.filteredRows[idx + 1]
                    visibleRows = self._getVisibleRows()
                    if idx + 1 >= self._filterViewIndex + visibleRows:
                        self._filterViewIndex = idx + 1 - visibleRows + 1
                    self._renderGrid()
            except ValueError:
                pass
        elif self.cursorRow < self.maxRow:
            self.cursorRow += 1
            visibleRows = self._getVisibleRows()
            if self.cursorRow >= self.viewTopRow + visibleRows:
                self.viewTopRow = self.cursorRow - visibleRows + 1
            self._renderGrid()

    def on_mouse_scroll_up(self, event) -> None:
        event.stop()
        if self.filteredRows is not None:
            try:
                idx = self.filteredRows.index(self.cursorRow)
                if idx > 0:
                    self.cursorRow = self.filteredRows[idx - 1]
                    if idx - 1 < self._filterViewIndex:
                        self._filterViewIndex = max(0, idx - 1)
                    self._renderGrid()
            except ValueError:
                pass
        elif self.cursorRow > 1:
            self.cursorRow -= 1
            if self.cursorRow < self.viewTopRow:
                self.viewTopRow = self.cursorRow
            self._renderGrid()

    def on_click(self, event) -> None:
        if self.isSearchMode or self.isFilterMode or self.isGotoRowMode:
            return
        try:
            gridWidget = self.query_one("#gridContent", Static)
            contentRegion = gridWidget.content_region
        except Exception:
            return
        contentX = event.screen_x - contentRegion.x
        contentY = event.screen_y - contentRegion.y
        if contentX < 0 or contentY < 0:
            return
        if contentX >= contentRegion.width or contentY >= contentRegion.height:
            return
        self._clickToCell(contentX, contentY)

    def _clickToCell(self, contentX: int, contentY: int) -> None:
        """根据 #gridContent 内容区坐标映射到 (row, col) 并移动光标"""
        rowNumWidth = max(7, displayWidth(str(self.maxRow)) + 2)
        colSepWidth = 3  # " │ " 占 3 个终端列

        if contentY == 0:
            clickedRow = 1
        else:
            rowsToRender = self._getRowsToRender()
            dataIdx = contentY - 1
            if dataIdx >= len(rowsToRender):
                return
            clickedRow = rowsToRender[dataIdx]

        if contentX < rowNumWidth:
            self.cursorRow = clickedRow
            self._renderGrid()
            return

        x = rowNumWidth
        clickedCol = self.viewLeftCol
        endCol = self._computeEndCol(rowNumWidth)
        for c in range(self.viewLeftCol, endCol + 1):
            if c > self.viewLeftCol:
                x += colSepWidth
            colW = self._getColWidth(c)
            if contentX < x + colW:
                clickedCol = c
                break
            x += colW
        else:
            clickedCol = endCol

        self.cursorRow = clickedRow
        self.cursorCol = max(1, min(clickedCol, self.maxCol))
        self._renderGrid()


# ---------- 行详情界面 ----------

class RowViewScreen(Screen):
    BINDINGS = [
        Binding("q", "back", "返回"),
        Binding("c", "copy_cell", "拷贝"),
    ]

    def __init__(
        self,
        workbook,
        sheetName: str,
        rowIndex: int,
        filePath: str,
        cursorCol: int = 1,
        **kwargs,
    ):
        super().__init__(**kwargs)
        self.workbook = workbook
        self.sheetName = sheetName
        self.rowIndex = rowIndex
        self.filePath = filePath
        self.ws = workbook[sheetName]
        self.maxCol = self.ws.max_column or 1
        self._initCursorCol = min(max(1, cursorCol), self.maxCol)

    def compose(self) -> ComposeResult:
        yield Header(show_clock=False)
        yield Static(
            f"[bold green]{Path(self.filePath).name}[/] [dim]→[/] [bold]{self.sheetName}[/] "
            f"[dim]→[/] [bold]行{self.rowIndex}[/]  [dim]↑↓选择 c拷贝 q返回[/]",
            id="rowBanner",
        )
        yield DataTable(id="rowTable", cursor_type="row", zebra_stripes=True)
        yield Footer()

    def on_mount(self) -> None:
        rows: list[tuple[str, str]] = []
        for c in range(1, self.maxCol + 1):
            fieldName = getCellValue(self.ws, 1, c) or get_column_letter(c)
            val = formatDisplayValue(getCellValue(self.ws, self.rowIndex, c))
            rows.append((fieldName, val))

        nameWidth = max((displayWidth(r[0]) for r in rows), default=4)
        nameWidth = max(nameWidth, displayWidth("字段名"))
        valWidth = max((displayWidth(r[1]) for r in rows), default=4)
        valWidth = max(valWidth, displayWidth("值"))

        table = self.query_one(DataTable)
        table.add_column("字段名", width=nameWidth)
        table.add_column("值", width=valWidth)
        for fieldName, val in rows:
            table.add_row(fieldName, val)
        table.move_cursor(row=self._initCursorCol - 1)
        table.focus()

    @property
    def cursorCol(self) -> int:
        try:
            return self.query_one(DataTable).cursor_row + 1
        except Exception:
            return 1

    def action_back(self) -> None:
        self.app.pop_screen()

    def action_copy_cell(self) -> None:
        val = getCellValue(self.ws, self.rowIndex, self.cursorCol)
        text = "" if val is None else str(val)
        try:
            self.app.copy_to_clipboard(text)
            self.notify("Copied to clipboard", timeout=2)
        except Exception as e:
            self.notify(f"Copy failed: {e}", severity="error", timeout=3)

    def key_c(self, event: Key) -> None:
        event.prevent_default()
        event.stop()
        self.action_copy_cell()

    def key_q(self, event: Key) -> None:
        event.prevent_default()
        event.stop()
        self.app.pop_screen()

    def key_escape(self, event: Key) -> None:
        event.prevent_default()
        event.stop()
        self.app.pop_screen()


# ---------- 应用入口 ----------

class ExcelTuiApp(App):
    TITLE = "Excel TUI"
    CSS = """
    #fileInfo { padding: 1; margin: 1; }
    #sheetTitle { padding: 1; margin: 1; }
    .sheetItem { padding: 0 1; margin: 0 1; }
    #gridContainer {
        height: 1fr;
        min-height: 20;
        padding: 0 2; border: solid $primary;
        overflow-x: auto; overflow-y: auto;
    }
    #gridContent { padding: 0 1; }
    #schemaPanel { padding: 0 0 0 4; height: auto; border: none; }
    #schemaEnRow { color: cyan; }
    #schemaTypeRow { color: yellow; }
    #rowTable { height: 1fr; }
    #sheetHeader, #rowBanner { padding: 0 1; }
    #searchInput, #filterInput, #gotoRowInput {
        dock: top; margin: 1; width: 50;
    }
    """

    def __init__(self, filePath: str, schemaData: dict | None = None, **kwargs):
        super().__init__(**kwargs)
        self.filePath = filePath
        self.schemaData = schemaData
        self.workbook = None

    def on_mount(self) -> None:
        from .workbook import loadWorkbook
        try:
            self.workbook = loadWorkbook(self.filePath)
        except Exception as e:
            self.notify(f"Failed to load file: {e}", severity="error")
            self.exit(1)
            return
        if len(self.workbook.sheetnames) == 1:
            sheetName = self.workbook.sheetnames[0]
            self.push_screen(
                SheetViewScreen(self.workbook, sheetName, self.filePath, schemaData=self.schemaData)
            )
        else:
            self.push_screen(
                SheetSelectScreen(self.workbook, self.filePath, schemaData=self.schemaData)
            )
