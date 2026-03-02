#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工作簿数据模型 + 多格式文件加载（xlsx/xls/csv）
"""

import csv
from pathlib import Path

from openpyxl import load_workbook

try:
    import xlrd
    HAS_XLS = True
except ImportError:
    HAS_XLS = False


# ---------- 统一 Sheet 数据结构 ----------

class _XlsCell:
    def __init__(self, sheet: "_XlsSheetWrapper", row: int, col: int) -> None:
        self._sheet = sheet
        self._row = row
        self._col = col

    @property
    def value(self) -> str:
        val = self._sheet._data.get((self._row, self._col), "")
        return "" if val is None else str(val)


class _XlsSheetWrapper:
    def __init__(self, name: str, data: dict[tuple[int, int], str]) -> None:
        self.title = name
        self._data = data
        self._maxRow = max((r for r, _ in data), default=1) if data else 1
        self._maxCol = max((c for _, c in data), default=1) if data else 1

    @property
    def max_row(self) -> int:
        return self._maxRow

    @property
    def max_column(self) -> int:
        return self._maxCol

    def cell(self, row: int, column: int) -> _XlsCell:
        return _XlsCell(self, row, column)


class _XlsWorkbook:
    def __init__(self, sheets: dict[str, _XlsSheetWrapper]) -> None:
        self._sheets = sheets
        self.sheetnames = list(sheets.keys())

    def __getitem__(self, name: str) -> _XlsSheetWrapper:
        return self._sheets[name]


# ---------- 格式加载 ----------

def _loadXls(path: str) -> _XlsWorkbook:
    """加载 xls，使用 row_values 批量读取加速"""
    if not HAS_XLS:
        raise RuntimeError("xls support requires xlrd: pip install xlrd")
    rd = xlrd.open_workbook(path)
    sheets = {}
    for i in range(rd.nsheets):
        sh = rd.sheet_by_index(i)
        data: dict[tuple[int, int], str] = {}
        for r in range(sh.nrows):
            rowVals = sh.row_values(r)
            for c in range(sh.ncols):
                val = rowVals[c] if c < len(rowVals) else ""
                data[(r + 1, c + 1)] = "" if val is None else str(val)
        if not data:
            data[(1, 1)] = ""
        sheets[sh.name] = _XlsSheetWrapper(sh.name, data)
    return _XlsWorkbook(sheets)


def _loadReadOnlyXlsx(path: str) -> _XlsWorkbook:
    """只读模式加载 xlsx，使用 read_only 流式读取大幅加速"""
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            data: dict[tuple[int, int], str] = {}
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    val = cell.value
                    data[(row_idx, col_idx)] = (
                        "" if val is None else str(val).strip().replace("\n", " ").replace("\r", " ")
                    )
            if not data:
                data[(1, 1)] = ""
            sheets[name] = _XlsSheetWrapper(name, data)
        return _XlsWorkbook(sheets)
    finally:
        wb.close()


def _loadCsv(path: str) -> _XlsWorkbook:
    """加载 CSV 文件，自动尝试 utf-8-sig / gbk 编码"""
    sheetName = Path(path).stem
    data: dict[tuple[int, int], str] = {}
    for encoding in ("utf-8-sig", "gbk", "latin-1"):
        try:
            with open(path, newline="", encoding=encoding) as f:
                reader = csv.reader(f)
                for rowIdx, row in enumerate(reader, start=1):
                    for colIdx, val in enumerate(row, start=1):
                        data[(rowIdx, colIdx)] = val.strip().replace("\n", " ").replace("\r", " ")
            break
        except (UnicodeDecodeError, UnicodeError):
            data = {}
            continue
    if not data:
        data[(1, 1)] = ""
    return _XlsWorkbook({sheetName: _XlsSheetWrapper(sheetName, data)})


def isXlsFile(path: str) -> bool:
    return Path(path).suffix.lower() == ".xls"


def loadWorkbook(path: str) -> _XlsWorkbook:
    """加载工作簿，所有格式均以只读模式打开"""
    suffix = Path(path).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _loadReadOnlyXlsx(path)
    if suffix == ".xls":
        return _loadXls(path)
    if suffix == ".csv":
        return _loadCsv(path)
    raise ValueError(f"Unsupported format: {suffix}")


def getCellValue(ws, row: int, col: int) -> str:
    try:
        val = ws.cell(row=row, column=col).value
        return "" if val is None else str(val)
    except Exception:
        return ""
